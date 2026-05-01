import re
import time
import random
import json
import subprocess
import requests
import sys
from pathlib import Path
from openpyxl.styles import PatternFill
import streamlit as st
import pandas as pd
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from io import BytesIO
from urllib.parse import urljoin

# Tự cài Playwright browser khi chạy trên cloud
@st.cache_resource
def install_playwright():
    # Chỉ cài browser binary — system libs đã được cài qua packages.txt
    subprocess.run(
        [sys.executable, "-m", "playwright", "install", "chromium"],
        check=False
    )

install_playwright()

ALN_DOMAIN = "https://alonhadat.com.vn"
BDS_DOMAIN = "https://batdongsan.com.vn"
NT_DOMAIN  = "https://www.nhatot.com"

# ─── Shared helpers ───────────────────────────────────────────────────────────
def clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()

def extract_loai_duong(text: str) -> str:
    t = text.lower()
    if re.search(r"m[aặ]t\s*ti[eề]n", t):
        return "mặt tiền"
    m = re.search(r"h[eẻ]m\s*(\d+[\w/]*)?", t)
    return clean(m.group(0)) if m else ""

# ─── alonhadat (Playwright) ───────────────────────────────────────────────────
# ====================== PARSE DETAIL ======================
def aln_parse_detail(soup: BeautifulSoup, url: str) -> dict:
    def txt(sel):
        el = soup.select_one(sel)
        return clean(el.get_text()) if el else ""

    tieu_de = txt("h1") or txt("header h1") or txt(".title")
    mo_ta   = txt("section.detail") or txt(".detail-content") or txt(".description")
    ngay    = txt("time, .date, .published-date").replace("Ngày đăng:", "").strip()
    gia     = txt(".price, span.price, .gia").replace("Giá:", "").strip()
    dt      = txt(".area, .dien-tich, span.area").replace("Diện tích:", "").strip()

    # Địa chỉ
    dia_chi = ""
    addr = soup.select_one("address, .address, .location")
    if addr:
        dia_chi = clean(addr.get_text())

    return {
        "nguon": "alonhadat",
        "tieu_de": tieu_de,
        "url": url,
        "ngay_dang": ngay,
        "gia": gia,
        "dien_tich": dt,
        "dia_chi": dia_chi,
        "loai_duong": extract_loai_duong(tieu_de + " " + mo_ta),
        "mo_ta": mo_ta[:300],
    }


STEALTH_JS = """
    Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
    Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
    Object.defineProperty(navigator, 'languages', {get: () => ['vi-VN', 'vi', 'en-US']});
    window.chrome = { runtime: {} };
"""

ALN_PROFILE_DIR = Path.home() / ".bds_scraper" / "aln_profile"
BDS_PROFILE_DIR = Path.home() / ".bds_scraper" / "bds_profile"
NT_PROFILE_DIR  = Path.home() / ".bds_scraper" / "nt_profile"

BLOCK_HANDLER = lambda r: (
    r.abort() if r.request.resource_type in ("image", "media", "font", "stylesheet")
    else r.continue_()
)

# Tín hiệu bot trong HTML source (bắt Cloudflare challenge page)
_BOT_HTML_SIGNALS = [
    "cf-turnstile",
    "cf-challenge",
    "challenge-form",
]
# Tín hiệu bot trong text hiển thị (bắt alonhadat và các trang dùng text thông báo)
_BOT_TEXT_SIGNALS = [
    "tôi không phải người máy",
    "kẻ xấu",
    "checking your browser",
    "verify you are human",
    "ddos-guard",
]

def _is_blocked(html: str) -> bool:
    return (any(sig in html for sig in _BOT_HTML_SIGNALS)
            or any(sig in html for sig in _BOT_TEXT_SIGNALS))

def _check_captcha(page, log, restore_handler=BLOCK_HANDLER) -> bool:
    """Dùng chung cho cả 3 trang.
    Python check: HTML source (bắt Cloudflare) + text (bắt alonhadat).
    JS wait: chỉ kiểm tra innerText — tránh false positive từ Cloudflare scripts thường trực trong <script>.
    """
    html = page.content().lower()
    if not _is_blocked(html):
        return True

    page.unroute("**/*")
    try:
        page.reload(wait_until="domcontentloaded", timeout=15000)
    except Exception:
        pass
    log("⚠️ Bot chặn! Nhìn vào cửa sổ trình duyệt và giải xác minh (tối đa 3 phút)...")

    # JS chỉ check innerText (text hiển thị) — KHÔNG check innerHTML
    # vì Cloudflare scripts luôn có trong HTML dù trang bình thường
    js_text_signals = json.dumps(_BOT_TEXT_SIGNALS)
    js_condition = f"""() => {{
        const text = document.body ? document.body.innerText.toLowerCase() : '';
        const signals = {js_text_signals};
        return !signals.some(s => text.includes(s)) && text.length > 500;
    }}"""

    try:
        page.wait_for_function(js_condition, timeout=180_000)
        page.wait_for_load_state("domcontentloaded", timeout=10_000)
        page.wait_for_timeout(1500)
        log("✓ Đã qua bot, tiếp tục scrape.")
    except Exception:
        log("✗ Hết thời gian chờ, dừng.")
        return False
    finally:
        page.route("**/*", restore_handler)

    return True

def scrape_alonhadat(base_url: str, num_pages: int, log, headless: bool = False) -> list[dict]:
    results = []
    launch_args = ["--no-sandbox", "--disable-blink-features=AutomationControlled",
                   "--disable-dev-shm-usage"]
    ctx_kwargs = dict(
        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        viewport={"width": 1366, "height": 768},
        locale="vi-VN",
        timezone_id="Asia/Ho_Chi_Minh",
    )

    with sync_playwright() as p:
        if headless:
            # Cloud: setup tối giản giống batdongsan (không stealth, không locale)
            browser = p.chromium.launch(headless=True, args=launch_args)
            context = browser.new_context(
                user_agent=ctx_kwargs["user_agent"],
                viewport=ctx_kwargs["viewport"],
            )
        else:
            # Local: persistent context để lưu session giải CAPTCHA
            ALN_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
            browser = None
            context = p.chromium.launch_persistent_context(
                str(ALN_PROFILE_DIR), headless=False, args=launch_args, **ctx_kwargs
            )
            context.add_init_script(STEALTH_JS)

        page = context.new_page()
        if not headless:
            page.route("**/*", lambda r: r.abort()
                if r.request.resource_type in ("image", "media", "font", "stylesheet")
                else r.continue_())

        for pg in range(1, num_pages + 1):
            if pg == 1:
                url = base_url
            else:
                base = base_url.rstrip("/")
                if base.endswith(".html"):
                    base = base[:-5]
                url = f"{base}/trang-{pg}.html"

            log(f"[alonhadat] trang {pg}: {url}")
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=30000)
                page.wait_for_timeout(2000)
            except Exception as e:
                log(f"[alonhadat] trang {pg}: lỗi load — {e}")
                continue

            if not _check_captcha(page, log):
                break

            soup = BeautifulSoup(page.content(), "html.parser")
            article_links = [
                urljoin(ALN_DOMAIN, a["href"])
                for art in soup.select("article.property-item")
                if (a := art.select_one("a.link")) and a.get("href")
            ]

            if not article_links:
                log(f"[alonhadat] trang {pg}: hết bài, dừng.")
                break

            log(f"[alonhadat] trang {pg}: {len(article_links)} bài")

            for link in article_links:
                try:
                    page.goto(link, wait_until="domcontentloaded", timeout=30000)
                    page.wait_for_timeout(1000)
                    if not _check_captcha(page, log):
                        break
                    item = aln_parse_detail(BeautifulSoup(page.content(), "html.parser"), link)
                    if item.get("tieu_de"):
                        results.append(item)
                        log(f"  ✓ {item['tieu_de'][:60]}")
                except Exception as e:
                    log(f"  ✗ lỗi chi tiết: {e}")
                time.sleep(random.uniform(1.0, 2.0))

            time.sleep(random.uniform(2.0, 3.5))

        context.close()
        if browser:
            browser.close()

    log(f"[alonhadat] xong — {len(results)} bài")
    return results

# ─── batdongsan (Playwright) ──────────────────────────────────────────────────
def bds_parse_cards(html: str) -> list[dict]:
    soup  = BeautifulSoup(html, "html.parser")
    items = []
    for card in soup.select("div.js__card"):
        a = card.select_one("a.js__product-link-for-product-id")
        if not a:
            continue
        href = a.get("href", "")
        if not href.startswith("http"):
            href = BDS_DOMAIN + href

        title_el = card.select_one("span.pr-title.js__card-title")
        price_el = (card.select_one("span.re__card-config-price.js__card-config-item")
                    or card.select_one("span.re__card-config-price"))
        area_el  = (card.select_one("span.re__card-config-area.js__card-config-item")
                    or card.select_one("span.re__card-config-area"))
        date_el  = card.select_one("span.re__card-published-info-published-at[aria-label]")
        loc_el   = card.select_one("div.re__card-location span")

        tieu_de = clean(title_el.get_text()) if title_el else ""
        dia_chi = clean(loc_el.get_text())   if loc_el   else ""
        if not tieu_de:
            continue

        items.append({
            "nguon": "batdongsan", "tieu_de": tieu_de, "url": href,
            "ngay_dang": date_el.get("aria-label", "") if date_el else "",
            "gia":        clean(price_el.get_text()) if price_el else "",
            "dien_tich":  clean(area_el.get_text())  if area_el  else "",
            "dia_chi":    dia_chi,
            "loai_duong": extract_loai_duong(tieu_de + " " + dia_chi),
            "mo_ta":      "",
        })
    return items

def scrape_batdongsan(base_url: str, num_pages: int, log) -> list[dict]:
    results = []
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage",
                  "--disable-blink-features=AutomationControlled"]
        )
        ctx  = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800},
        )
        ctx.add_init_script(STEALTH_JS)
        page = ctx.new_page()
        # Không block gì cả — BDS cần JS + CSS để render card trên cloud
        page.route("**/*", lambda r: r.abort()
            if r.request.resource_type == "media"
            else r.continue_())

        for pg in range(1, num_pages + 1):
            url = base_url if pg == 1 else f"{base_url.rstrip('/')}/p{pg}"
            log(f"[batdongsan] trang {pg}: {url}")
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=25000)
                # Chờ JS render xong rồi mới đọc HTML
                try:
                    page.wait_for_selector("div.js__card", timeout=8000)
                except Exception:
                    pass
                items = bds_parse_cards(page.content())
            except Exception as e:
                log(f"[batdongsan] trang {pg}: lỗi — {e}")
                continue

            if not items:
                log(f"[batdongsan] trang {pg}: hết bài, dừng.")
                break

            results.extend(items)
            log(f"[batdongsan] trang {pg}: {len(items)} bài")
            time.sleep(1.0)

        browser.close()
    return results

# ─── nhatot (Playwright) ──────────────────────────────────────────────────────
def nt_parse_cards(html: str) -> list[dict]:
    soup  = BeautifulSoup(html, "html.parser")
    items = []

    cards = soup.select("li[class*='adItem'], li[class*='AdItem']")
    if not cards:
        cards = [el for el in soup.select("li, article")
                 if el.select_one("h2") and el.select_one("a[href]")]

    for card in cards:
        try:
            a    = card.select_one("a[href]")
            href = a["href"] if a else ""
            if href and not href.startswith("http"):
                href = NT_DOMAIN + href

            title_el = card.select_one("h2, h3")
            tieu_de  = clean(title_el.get_text()) if title_el else ""
            if not tieu_de:
                continue

            gia = ""
            for span in card.select("span"):
                t = span.get_text(strip=True)
                if re.search(r"\d", t) and ("tỷ" in t or "triệu" in t) and "m²" not in t:
                    gia = clean(t); break

            dien_tich = ""
            for span in card.select("span"):
                t = span.get_text(strip=True)
                if "m²" in t and re.search(r"^\d", t) and "/" not in t:
                    dien_tich = clean(t); break

            dia_chi = ""
            for s in card.select("span"):
                if any(kw in s.get_text().lower() for kw in ("phường", "quận", "huyện", "tp")):
                    dia_chi = clean(s.get_text())
                    break

            items.append({
                "nguon": "nhatot", "tieu_de": tieu_de, "url": href,
                "ngay_dang": "", "gia": gia, "dien_tich": dien_tich,
                "dia_chi": dia_chi,
                "loai_duong": extract_loai_duong(tieu_de + " " + dia_chi),
                "mo_ta": "",
            })
        except Exception:
            continue
    return items

def scrape_nhatot(base_url: str, num_pages: int, log) -> list[dict]:
    results = []
    sep = "&" if "?" in base_url else "?"
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "vi-VN,vi;q=0.9,en;q=0.8",
        "Referer": NT_DOMAIN,
    })

    for pg in range(1, num_pages + 1):
        url = base_url if pg == 1 else f"{base_url}{sep}page={pg}"
        log(f"[nhatot] trang {pg}: {url}")
        try:
            r = session.get(url, timeout=15)
            r.raise_for_status()
            items = nt_parse_cards(r.text)
        except Exception as e:
            log(f"[nhatot] trang {pg}: lỗi — {e}")
            continue

        if not items:
            log(f"[nhatot] trang {pg}: hết bài, dừng.")
            break

        results.extend(items)
        log(f"[nhatot] trang {pg}: {len(items)} bài")
        time.sleep(1.0)

    return results

# ─── Excel export ─────────────────────────────────────────────────────────────
DONGIA_XAY_DUNG = 5  # triệu/m²

def extract_so_tang(text: str) -> int:
    t = text.lower()
    m = re.search(r'(\d+)\s*tầng', t)
    if m:
        return int(m.group(1))
    so_tret = 1 if 'trệt' in t else 0
    m_lau = re.search(r'(\d+)\s*lầu', t)
    if m_lau:
        return so_tret + int(m_lau.group(1))
    if 'lầu' in t:
        return so_tret + 1
    return 0

def parse_gia_trieu(s: str) -> float:
    s = s.lower().replace(',', '.')
    m = re.search(r'([\d.]+)\s*tỷ', s)
    if m:
        return round(float(m.group(1)) * 1000, 1)
    m = re.search(r'([\d.]+)\s*triệu', s)
    if m:
        return round(float(m.group(1)), 1)
    return 0.0

def parse_dien_tich_m2(s: str) -> float:
    m = re.search(r'([\d.,]+)', s.replace(',', '.'))
    return float(m.group(1)) if m else 0.0

def tinh_toan(df: pd.DataFrame) -> pd.DataFrame:
    gia_nha_col, don_gia_col = [], []
    for _, row in df.iterrows():
        nguon   = str(row.get("nguon", ""))
        tieu_de = str(row.get("tieu_de", ""))
        mo_ta   = str(row.get("mo_ta", "")) if nguon == "alonhadat" else ""

        so_tang = extract_so_tang(tieu_de)
        if so_tang == 0 and mo_ta:
            so_tang = extract_so_tang(mo_ta)

        dt  = parse_dien_tich_m2(str(row.get("dien_tich", "")))
        gia = parse_gia_trieu(str(row.get("gia", "")))

        if dt > 0 and gia > 0:
            gia_nha = dt * so_tang * DONGIA_XAY_DUNG  # = 0 nếu so_tang = 0
            gia_dat = gia - gia_nha
            don_gia = round(gia_dat / dt, 1)
            gia_nha_col.append(round(gia_nha, 1))
            don_gia_col.append(don_gia)
        else:
            gia_nha_col.append(None)
            don_gia_col.append(None)

    df["gia_nha (tr)"]       = gia_nha_col
    df["don_gia_dat (tr/m²)"] = don_gia_col

    # Sắp xếp theo diện tích tăng dần
    df["_dt_sort"] = df["dien_tich"].apply(parse_dien_tich_m2)
    df = df.sort_values("_dt_sort").drop(columns=["_dt_sort"]).reset_index(drop=True)
    return df

def to_excel(df: pd.DataFrame) -> bytes:
    buf = BytesIO()
    red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Kết quả")
        ws = writer.sheets["Kết quả"]
        col_don_gia = (list(df.columns).index("don_gia_dat (tr/m²)") + 1
                       if "don_gia_dat (tr/m²)" in df.columns else None)
        if col_don_gia:
            for r in range(2, len(df) + 2):
                cell = ws.cell(row=r, column=col_don_gia)
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    # tô đỏ toàn bộ hàng
                    for c in range(1, len(df.columns) + 1):
                        ws.cell(row=r, column=c).fill = red
    return buf.getvalue()

# ─── Streamlit UI ─────────────────────────────────────────────────────────────
st.set_page_config(page_title="BDS Scraper", layout="wide", page_icon="🏠")
st.title("🏠 BDS Scraper")
st.caption("Dán link danh sách từ mỗi trang vào ô tương ứng rồi nhấn Scrape")

col1, col2, col3 = st.columns(3)
with col1:
    url_aln = st.text_input("alonhadat.com.vn", placeholder="https://alonhadat.com.vn/...")
with col2:
    url_bds = st.text_input("batdongsan.com.vn", placeholder="https://batdongsan.com.vn/...")
with col3:
    url_nt  = st.text_input("nhatot.com", placeholder="https://www.nhatot.com/...")

num_pages = st.slider("Số trang cần scrape", 1, 20, 3)
cloud_mode = st.toggle(
    "☁️ Chế độ Cloud (tắt khi chạy local để giải CAPTCHA alonhadat)",
    value=True,
    help="Bật khi deploy trên Streamlit Cloud. Tắt khi chạy máy tính cá nhân để dùng browser thật giải CAPTCHA."
)

if st.button("🚀 Scrape", use_container_width=True):
    if not any([url_aln, url_bds, url_nt]):
        st.warning("Vui lòng dán ít nhất 1 link.")
    else:
        results  = []
        log_box  = st.empty()
        logs: list[str] = []

        def log(msg: str):
            logs.append(msg)
            log_box.code("\n".join(logs[-20:]))

        if url_aln:
            st.write("📥 Đang scrape alonhadat…")
            try:
                rows = scrape_alonhadat(url_aln.strip(), num_pages, log, headless=cloud_mode)
                results.extend(rows)
                log(f"[alonhadat] xong — {len(rows)} bài")
            except Exception as e:
                st.warning(f"alonhadat lỗi: {e}")

        if url_bds:
            st.write("📥 Đang scrape batdongsan…")
            try:
                rows = scrape_batdongsan(url_bds.strip(), num_pages, log)
                results.extend(rows)
                log(f"[batdongsan] xong — {len(rows)} bài")
            except Exception as e:
                st.warning(f"batdongsan lỗi: {e}")

        if url_nt:
            st.write("📥 Đang scrape nhatot…")
            try:
                rows = scrape_nhatot(url_nt.strip(), num_pages, log)
                results.extend(rows)
                log(f"[nhatot] xong — {len(rows)} bài")
            except Exception as e:
                st.warning(f"nhatot lỗi: {e}")

        if results:
            df = pd.DataFrame(results)
            df = tinh_toan(df)
            st.success(f"Tổng cộng {len(df)} bài đăng")

            show_cols = ["nguon", "tieu_de", "gia", "dien_tich", "dia_chi",
                         "loai_duong", "gia_nha (tr)", "don_gia_dat (tr/m²)"]
            show_cols = [c for c in show_cols if c in df.columns]

            styler = df[show_cols].style
            if "don_gia_dat (tr/m²)" in show_cols:
                styler = styler.map(
                    lambda v: "background-color: #FF9999" if isinstance(v, (int, float)) and v < 0 else "",
                    subset=["don_gia_dat (tr/m²)"]
                )
            st.dataframe(styler, use_container_width=True)
            st.download_button(
                "📥 Tải Excel",
                data=to_excel(df),
                file_name="bds.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("Không tìm thấy bài đăng nào.")
